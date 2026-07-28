import re
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction
from django.utils import timezone

from accounts.models import User
from user.models import LearnerProfile
from admin_dashboard.models import Centre

REQUIRED_HEADERS = ['Email', 'Contact', 'NIELIT Centre', 'Batch Code', 'Full Name', 'Gender']
VALID_GENDERS = {'male': 'male', 'female': 'female', 'other': 'other', 'm': 'male', 'f': 'female'}
EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
CONTACT_RE = re.compile(r'^\d{10}$')


def generate_common_password():
    """NIELIT<currentyear>delhi123 — same shared password for every bulk-created account."""
    return f"NIELIT{timezone.now().year}delhi123"


def build_upload_template():
    """Generates the downloadable .xlsx template with headers, an example row,
    an instructions sheet, and a reference list of exact active Centre names."""
    wb = Workbook()

    # ── Sheet 1: the actual data-entry sheet ──
    data_ws = wb.active
    data_ws.title = "Users"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="172281", end_color="172281", fill_type="solid")

    for col_idx, header in enumerate(REQUIRED_HEADERS, start=1):
        cell = data_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    example_row = ["student@example.com", "9876543210", "NIELIT Delhi Centre", "BATCH2026A", "Ravi Kumar", "Male"]
    for col_idx, value in enumerate(example_row, start=1):
        data_ws.cell(row=2, column=col_idx, value=value)

    for col_idx in range(1, len(REQUIRED_HEADERS) + 1):
        data_ws.column_dimensions[chr(64 + col_idx)].width = 26

    # ── Sheet 2: instructions ──
    instr_ws = wb.create_sheet("Instructions")
    instructions = [
        "INSTRUCTIONS — Bulk Student Upload",
        "",
        "1. Do not change the column headers in the 'Users' sheet.",
        "2. Row 2 is an example — replace it with real data, or delete it before uploading.",
        "3. All six fields are mandatory for every row: Email, Contact, NIELIT Centre, Batch Code, Full Name, Gender.",
        "4. Email must be a valid, unique email address — not already registered.",
        "5. Contact must be a unique 10-digit number.",
        "6. NIELIT Centre must exactly match a centre name already created in the admin dashboard — see the 'Valid Centres' sheet for the exact spelling to copy.",
        "7. Batch Code can be any text — it will automatically be converted to uppercase with spaces removed (e.g. 'batch 2026 a' becomes 'BATCH2026A').",
        "8. Gender must be Male, Female, or Other.",
        "9. If ANY row fails validation, NO accounts will be created — fix all listed errors and re-upload.",
        "10. All created accounts share the same initial password, which is emailed to each student individually along with their login email/contact.",
    ]
    for row_idx, line in enumerate(instructions, start=1):
        cell = instr_ws.cell(row=row_idx, column=1, value=line)
        if row_idx == 1:
            cell.font = Font(bold=True, size=13)
    instr_ws.column_dimensions['A'].width = 100

    # ── Sheet 3: exact valid centre names, for copy-paste accuracy ──
    centre_ws = wb.create_sheet("Valid Centres")
    centre_ws.cell(row=1, column=1, value="Centre Name").font = Font(bold=True)
    for row_idx, centre in enumerate(Centre.objects.filter(is_active=True).order_by('centre_name'), start=2):
        centre_ws.cell(row=row_idx, column=1, value=centre.centre_name)
    centre_ws.column_dimensions['A'].width = 40

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def parse_and_validate_upload(excel_file):
    """
    Reads the uploaded workbook and validates every row against DB state and
    format rules. Returns (valid_rows, errors) — valid_rows is only non-empty
    if errors is completely empty (all-or-nothing validation).
    """
    errors = []
    valid_rows = []

    try:
        wb = load_workbook(excel_file, data_only=True)
    except Exception:
        return [], [{'row': '-', 'message': "Could not read the file. Make sure it's a valid .xlsx file."}]

    if "Users" in wb.sheetnames:
        ws = wb["Users"]
    else:
        ws = wb.active

    header_row = [cell.value for cell in ws[1]]
    normalized_headers = [str(h).strip() if h else '' for h in header_row]

    if normalized_headers[:len(REQUIRED_HEADERS)] != REQUIRED_HEADERS:
        return [], [{
            'row': 1,
            'message': f"Header row must exactly match: {', '.join(REQUIRED_HEADERS)}. "
                       f"Found: {', '.join([h for h in normalized_headers if h]) or 'empty row'}",
        }]

    existing_emails = set(User.objects.values_list('email', flat=True))
    existing_contacts = set(User.objects.exclude(contact__isnull=True).values_list('contact', flat=True))
    active_centres = {c.centre_name.strip().lower(): c for c in Centre.objects.filter(is_active=True)}

    seen_emails_in_file = set()
    seen_contacts_in_file = set()

    row_num = 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None or str(v).strip() == '' for v in row):
            row_num += 1
            continue   # skip fully blank rows silently

        email, contact, centre_name, batch_code, full_name, gender = (
            (row + (None,) * 6)[:6]   # pad in case a row has fewer cells than expected
        )

        row_errors = []

        email = str(email).strip().lower() if email else ''
        if not email:
            row_errors.append("Email is required.")
        elif not EMAIL_RE.match(email):
            row_errors.append("Email format is invalid.")
        elif email in existing_emails:
            row_errors.append(f"Email '{email}' is already registered.")
        elif email in seen_emails_in_file:
            row_errors.append(f"Email '{email}' is duplicated within this file.")

        contact = str(contact).strip() if contact else ''
        contact = re.sub(r'\D', '', contact)   # strip any non-digits (spaces, dashes, +91 etc.)
        if not contact:
            row_errors.append("Contact is required.")
        elif not CONTACT_RE.match(contact):
            row_errors.append("Contact must be exactly 10 digits.")
        elif contact in existing_contacts:
            row_errors.append(f"Contact '{contact}' is already registered.")
        elif contact in seen_contacts_in_file:
            row_errors.append(f"Contact '{contact}' is duplicated within this file.")

        centre_name_clean = str(centre_name).strip() if centre_name else ''
        centre_obj = active_centres.get(centre_name_clean.lower())
        if not centre_name_clean:
            row_errors.append("NIELIT Centre is required.")
        elif not centre_obj:
            row_errors.append(f"'{centre_name_clean}' does not match any active centre. Check the 'Valid Centres' sheet for exact names.")

        batch_code_clean = str(batch_code).strip().upper().replace(' ', '') if batch_code else ''
        if not batch_code_clean:
            row_errors.append("Batch Code is required.")

        full_name_clean = str(full_name).strip() if full_name else ''
        if not full_name_clean:
            row_errors.append("Full Name is required.")

        gender_clean = str(gender).strip().lower() if gender else ''
        gender_value = VALID_GENDERS.get(gender_clean)
        if not gender_clean:
            row_errors.append("Gender is required.")
        elif not gender_value:
            row_errors.append(f"Gender '{gender}' is invalid. Use Male, Female, or Other.")

        if row_errors:
            errors.append({'row': row_num, 'message': '; '.join(row_errors)})
        else:
            seen_emails_in_file.add(email)
            seen_contacts_in_file.add(contact)
            valid_rows.append({
                'email': email,
                'contact': contact,
                'centre': centre_obj,
                'batch_code': batch_code_clean,
                'full_name': full_name_clean,
                'gender': gender_value,
            })

        row_num += 1

    if not valid_rows and not errors:
        errors.append({'row': '-', 'message': "No data rows found below the header row."})

    if errors:
        return [], errors   # all-or-nothing: any error blocks every row, even valid ones
    return valid_rows, []


@transaction.atomic
def create_users_from_rows(valid_rows, created_by):
    """Creates User + LearnerProfile for every validated row. Returns the created User list."""
    common_password = generate_common_password()
    created_users = []

    for row in valid_rows:
        user = User(
            email=row['email'],
            contact=row['contact'],
            role=User.Role.USER,
            nielit_centre=row['centre'],
            batch_code=row['batch_code'],
            account_status=User.AccountStatus.ACTIVE,   # admin-uploaded rows are pre-vetted — active immediately
            is_active=True,
            account_status_updated_at=timezone.now(),
        )
        user.set_password(common_password)
        user.save()

        LearnerProfile.objects.create(
            user=user,
            full_name=row['full_name'],
            gender=row['gender'],
            profile_completed=False,   # still must complete the rest of their profile
        )
        created_users.append(user)

    return created_users, common_password